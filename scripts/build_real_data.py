#!/usr/bin/env python3
"""Build the static release from official public source archives.

This script intentionally does not invent missing relationships. It keeps industry
taxonomies and occupation/work data as separate evidence-backed collections.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
OUT = ROOT / "client" / "public" / "data"
OUT.mkdir(parents=True, exist_ok=True)


SOURCES = {
    "onet": {
        "path": RAW / "onet_30_3_csv.zip",
        "publisher": "U.S. Department of Labor / O*NET Resource Center",
        "title": "O*NET 30.3 Database",
        "vintage": "30.3",
        "url": "https://www.onetcenter.org/database.html",
        "license": "CC BY 4.0",
    },
    "bls": {
        "path": RAW / "bls_oesm25nat.zip",
        "publisher": "U.S. Bureau of Labor Statistics",
        "title": "May 2025 Occupational Employment and Wage Statistics — National",
        "vintage": "May 2025",
        "url": "https://www.bls.gov/oes/tables.htm",
        "license": "U.S. government public data; verify source notes before redistribution",
    },
    "blsProjections": {
        "path": RAW / "bls_ep_2024_2034_occupation.xlsx",
        "publisher": "U.S. Bureau of Labor Statistics",
        "title": "Occupational projections and worker characteristics, 2024–2034",
        "vintage": "2024–2034 (released 2025-08-28)",
        "url": "https://www.bls.gov/emp/tables/occupational-projections-and-characteristics.htm",
        "license": "U.S. government public data; projections are national estimates with documented uncertainty",
    },
    "naics": {
        "path": RAW / "naics_2022_structure.xlsx",
        "publisher": "U.S. Census Bureau",
        "title": "2022 NAICS Structure",
        "vintage": "2022",
        "url": "https://www.census.gov/naics/",
        "license": "U.S. government public data; retain attribution",
    },
    "isic": {
        "path": RAW / "isic_rev5_structure.csv",
        "publisher": "United Nations Statistics Division",
        "title": "ISIC Rev. 5 English Structure",
        "vintage": "Rev. 5",
        "url": "https://unstats.un.org/unsd/classifications/Econ/isic",
        "license": "UN source; retain attribution and source terms",
    },
    "isco": {
        "path": RAW / "isco_08_structure.xlsx",
        "publisher": "International Labour Organization / ILOSTAT",
        "title": "ISCO-08 Structure and Definitions",
        "vintage": "ISCO-08",
        "url": "https://ilostat.ilo.org/methods/concepts-and-definitions/classification-occupation/",
        "license": "Free to use without prior authorization; retain source attribution",
    },
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(name: str, value: object) -> None:
    (OUT / name).write_text(json.dumps(value, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def csv_rows(archive: zipfile.ZipFile, filename: str):
    member = next(name for name in archive.namelist() if name.endswith(f"/{filename}"))
    raw = archive.read(member)
    return csv.DictReader(io.TextIOWrapper(io.BytesIO(raw), encoding="utf-8-sig", newline=""))


def clean_number(value):
    if value in (None, "", "*", "**", "#"):
        return None
    try:
        number = float(str(value).replace(",", ""))
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return None


def parse_onet() -> tuple[list[dict], dict]:
    """Parse only direct O*NET facts; personal readiness and career ladders stay out of the release."""
    archive_path = SOURCES["onet"]["path"]
    with zipfile.ZipFile(archive_path) as archive:
        occupation_rows = list(csv_rows(archive, "occupation_data.csv"))
        tasks_by_code: dict[str, list[dict]] = defaultdict(list)
        for row in csv_rows(archive, "task_statements.csv"):
            code = row["O*NET-SOC Code"]
            tasks_by_code[code].append({"id": row["Task ID"], "text": row["Task"], "type": row["Task Type"], "date": row["Date"]})

        skills_by_code: dict[str, list[dict]] = defaultdict(list)
        for row in csv_rows(archive, "essential_skills.csv"):
            if row["Scale ID"] != "IM":
                continue
            skills_by_code[row["O*NET-SOC Code"]].append({"name": row["Element Name"], "importance": clean_number(row["Data Value"]), "date": row["Date"]})

        activities_by_code: dict[str, set[str]] = defaultdict(set)
        for row in csv_rows(archive, "work_activities.csv"):
            if row["Scale ID"] == "IM":
                activities_by_code[row["O*NET-SOC Code"]].add(row["Element Name"])

        software_by_code: dict[str, set[str]] = defaultdict(set)
        for row in csv_rows(archive, "software_skills.csv"):
            software_by_code[row["O*NET-SOC Code"]].add(row["Element Name"])

        titles_by_code: dict[str, set[str]] = defaultdict(set)
        for row in csv_rows(archive, "job_titles.csv"):
            titles_by_code[row["O*NET-SOC Code"]].add(row["Job Title"])

        job_zone_by_code: dict[str, dict] = {}
        for row in csv_rows(archive, "job_zones.csv"):
            job_zone_by_code[row["O*NET-SOC Code"]] = {"number": clean_number(row["Job Zone"]), "date": row["Date"]}

        zone_reference: dict[int, dict] = {}
        for row in csv_rows(archive, "job_zone_reference.csv"):
            for number in re.findall(r"\d+", row["Job Zone"]):
                zone_reference[int(number)] = {
                    "label": row["Job Zone"],
                    "experience": row["Experience"].replace("\n", " ").strip(),
                    "education": row["Education"].replace("\n", " ").strip(),
                    "training": row["Job Training"].replace("\n", " ").strip(),
                }

        education_categories = {
            row["Category"]: row["Category Description"]
            for row in csv_rows(archive, "education_categories.csv")
            if row["Element ID"] == "2.D.1"
        }
        education_by_code: dict[str, list[dict]] = defaultdict(list)
        for row in csv_rows(archive, "education.csv"):
            if row["Element ID"] != "2.D.1" or row["Scale ID"] != "RL":
                continue
            value = clean_number(row["Data Value"])
            if value is not None:
                education_by_code[row["O*NET-SOC Code"]].append({"label": education_categories.get(row["Category"], row["Category"]), "share": value, "date": row["Date"]})

        training_categories = {
            (row["Element ID"], row["Category"]): row["Category Description"]
            for row in csv_rows(archive, "training_and_experience_categories.csv")
        }
        preparation_by_code: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
        preparation_types = {"3.A.1": "relatedExperience", "3.A.2": "onSiteTraining", "3.A.3": "apprenticeshipTraining"}
        for row in csv_rows(archive, "training_and_experience.csv"):
            item_type = preparation_types.get(row["Element ID"])
            if not item_type:
                continue
            value = clean_number(row["Data Value"])
            if value is not None:
                preparation_by_code[row["O*NET-SOC Code"]][item_type].append({"label": training_categories.get((row["Element ID"], row["Category"]), row["Category"]), "share": value, "date": row["Date"]})

        related_by_code: dict[str, list[dict]] = defaultdict(list)
        for row in csv_rows(archive, "related_occupations.csv"):
            related_by_code[row["O*NET-SOC Code"]].append({
                "id": row["Related O*NET-SOC Code"],
                "title": row["Related Title"],
                "tier": row["Relatedness Tier"],
                "index": clean_number(row["Index"]),
            })

    occupations = []
    for row in occupation_rows:
        code = row["O*NET-SOC Code"]
        task_rows = tasks_by_code.get(code, [])
        skill_rows = sorted(skills_by_code.get(code, []), key=lambda item: item["importance"] or 0, reverse=True)
        zone = job_zone_by_code.get(code)
        zone_number = int(zone["number"]) if zone and zone.get("number") is not None else None
        preparation = {
            "jobZone": {**zone, **zone_reference.get(zone_number, {})} if zone and zone_number is not None else None,
            "reportedEducation": max(education_by_code.get(code, []), key=lambda item: item["share"], default=None),
            "relatedExperience": max(preparation_by_code.get(code, {}).get("relatedExperience", []), key=lambda item: item["share"], default=None),
            "onSiteTraining": max(preparation_by_code.get(code, {}).get("onSiteTraining", []), key=lambda item: item["share"], default=None),
            "apprenticeshipTraining": max(preparation_by_code.get(code, {}).get("apprenticeshipTraining", []), key=lambda item: item["share"], default=None),
        }
        occupation = {
            "id": code,
            "soc": code.split(".")[0],
            "title": row["Title"],
            "description": row["Description"],
            "level": "occupation",
            "source": {"id": "onet", "publisher": SOURCES["onet"]["publisher"], "vintage": SOURCES["onet"]["vintage"], "url": SOURCES["onet"]["url"]},
            "metrics": {
                "taskCount": len(task_rows),
                "skillCount": len(skill_rows),
                "workActivityCount": len(activities_by_code.get(code, set())),
                "softwareCount": len(software_by_code.get(code, set())),
                "alternateTitleCount": len(titles_by_code.get(code, set())),
            },
            "preparation": preparation,
            "relatedOccupations": sorted(related_by_code.get(code, []), key=lambda item: (item["index"] or 999, item["title"]))[:8],
            "tasks": task_rows[:10],
            "skills": skill_rows[:12],
            "workActivities": sorted(activities_by_code.get(code, set()))[:20],
            "software": sorted(software_by_code.get(code, set()))[:20],
            "alternateTitles": sorted(titles_by_code.get(code, set()))[:20],
        }
        occupations.append(occupation)

    occupations.sort(key=lambda item: item["title"].lower())
    return occupations, {
        "occupationCount": len(occupations),
        "taskStatementCount": sum(item["metrics"]["taskCount"] for item in occupations),
        "occupationSkillLinkCount": sum(item["metrics"]["skillCount"] for item in occupations),
        "workActivityLinkCount": sum(item["metrics"]["workActivityCount"] for item in occupations),
        "relatedOccupationLinkCount": sum(len(item["relatedOccupations"]) for item in occupations),
    }


def parse_bls() -> dict[str, dict]:
    archive_path = SOURCES["bls"]["path"]
    with zipfile.ZipFile(archive_path) as archive:
        member = next(name for name in archive.namelist() if name.endswith("national_M2025_dl.xlsx"))
        workbook = openpyxl.load_workbook(io.BytesIO(archive.read(member)), read_only=True, data_only=True)
        sheet = workbook["national_M2025_dl"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        index = {str(value): position for position, value in enumerate(header) if value is not None}
        result = {}
        for row in rows:
            occ_code = str(row[index["OCC_CODE"]] or "").strip()
            if not re.fullmatch(r"\d{2}-\d{4}", occ_code) or row[index["O_GROUP"]] != "detailed":
                continue
            if str(row[index["NAICS"]] or "") not in {"000000", "0", "0.0"}:
                continue
            result[occ_code] = {
                "title": row[index["OCC_TITLE"]],
                "employment": clean_number(row[index["TOT_EMP"]]),
                "employmentPerThousand": clean_number(row[index["JOBS_1000"]]),
                "locationQuotient": clean_number(row[index["LOC_QUOTIENT"]]),
                "medianAnnualWage": clean_number(row[index["A_MEDIAN"]]),
                "meanAnnualWage": clean_number(row[index["A_MEAN"]]),
                "source": {"id": "bls", "publisher": SOURCES["bls"]["publisher"], "vintage": SOURCES["bls"]["vintage"], "url": SOURCES["bls"]["url"]},
            }
        return result


def parse_bls_projections() -> dict[str, dict]:
    """Return published detailed National Employment Matrix records keyed by exact SOC code."""
    workbook = openpyxl.load_workbook(SOURCES["blsProjections"]["path"], read_only=True, data_only=True)
    sheet = workbook["Table 1.2"]
    rows = sheet.iter_rows(values_only=True)
    next(rows)  # table title
    header = next(rows)
    index = {str(value): position for position, value in enumerate(header) if value is not None}
    result: dict[str, dict] = {}
    for row in rows:
        code = code_as_string(row[index["2024 National Employment Matrix code"]])
        kind = code_as_string(row[index["Occupation type"]])
        if not re.fullmatch(r"\d{2}-\d{4}", code) or kind != "Line item":
            continue
        result[code] = {
            "title": code_as_string(row[index["2024 National Employment Matrix title"]]).strip(),
            "employment2024Thousands": clean_number(row[index["Employment, 2024"]]),
            "employment2034Thousands": clean_number(row[index["Employment, 2034"]]),
            "changeThousands": clean_number(row[index["Employment change, numeric, 2024–34"]]),
            "changePercent": clean_number(row[index["Employment change, percent, 2024–34"]]),
            "annualOpeningsThousands": clean_number(row[index["Occupational openings, 2024–34 annual average"]]),
            "source": {"id": "bls-ep", "publisher": SOURCES["blsProjections"]["publisher"], "vintage": SOURCES["blsProjections"]["vintage"], "url": SOURCES["blsProjections"]["url"]},
        }
    return result


def code_as_string(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_naics() -> list[dict]:
    workbook = openpyxl.load_workbook(SOURCES["naics"]["path"], read_only=True, data_only=True)
    sheet = workbook.active
    items = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        code = code_as_string(row[1])
        title = code_as_string(row[2]).strip()
        if not code or not title or not re.search(r"\d", code):
            continue
        title = re.sub(r"[T*]+\s*$", "", title).strip()
        level = len(code.split("-")[0]) if "-" in code else len(code)
        items.append({"id": f"naics:{code}", "code": code, "title": title, "level": level, "taxonomy": "NAICS 2022", "source": {"id": "naics", "publisher": SOURCES["naics"]["publisher"], "vintage": SOURCES["naics"]["vintage"], "url": SOURCES["naics"]["url"]}})
    return items


def parse_isic() -> list[dict]:
    raw = SOURCES["isic"]["path"].read_text(encoding="cp1252", errors="replace")
    reader = csv.reader(io.StringIO(raw))
    items = []
    for row in reader:
        if len(row) < 2:
            continue
        code = row[0].strip().strip('"')
        title = row[1].strip().strip('"')
        if not code or code.lower() in {"isic", "code", "isic rev 5 code"} or not re.fullmatch(r"[A-Z]|\d{2,4}", code):
            continue
        level = 1 if code.isalpha() else len(code)
        items.append({"id": f"isic:{code}", "code": code, "title": title, "level": level, "taxonomy": "ISIC Rev. 5", "source": {"id": "isic", "publisher": SOURCES["isic"]["publisher"], "vintage": SOURCES["isic"]["vintage"], "url": SOURCES["isic"]["url"]}})
    return items


def parse_isco() -> list[dict]:
    workbook = openpyxl.load_workbook(SOURCES["isco"]["path"], read_only=True, data_only=True)
    sheet = workbook["ISCO-08 EN Struct and defin"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    index = {str(value): position for position, value in enumerate(header) if value is not None}
    items = []
    for row in rows:
        code = code_as_string(row[index["ISCO 08 Code"]])
        title = code_as_string(row[index["Title EN"]]).strip()
        level = clean_number(row[index["Level"]])
        if not code or not title or level not in {1, 2, 3, 4} or not re.fullmatch(r"\d{1,4}", code):
            continue
        items.append({
            "id": f"isco:{code}", "code": code, "title": title, "level": int(level), "taxonomy": "ISCO-08",
            "definition": code_as_string(row[index["Definition"]]).replace("\n", " ").strip(),
            "source": {"id": "isco", "publisher": SOURCES["isco"]["publisher"], "vintage": SOURCES["isco"]["vintage"], "url": SOURCES["isco"]["url"]},
        })
    return items


def main() -> None:
    missing = [key for key, source in SOURCES.items() if not source["path"].exists()]
    if missing:
        raise SystemExit(f"Missing source files: {', '.join(missing)}")

    occupations, onet_counts = parse_onet()
    bls = parse_bls()
    projections = parse_bls_projections()
    for occupation in occupations:
        occupation["laborMarket"] = bls.get(occupation["soc"])
        occupation["outlook"] = projections.get(occupation["soc"])

    naics = parse_naics()
    isic = parse_isic()
    isco = parse_isco()
    taxonomies = naics + isic + isco
    counts = {
        "naics": len(naics),
        "isic": len(isic),
        "isco": len(isco),
        "occupations": len(occupations),
        "occupationsWithBls": sum(1 for item in occupations if item["laborMarket"]),
        "occupationsWithBlsProjections": sum(1 for item in occupations if item["outlook"]),
        "blsDetailedRecords": len(bls),
        "blsProjectionDetailedRecords": len(projections),
    }
    release_id = datetime.now(timezone.utc).strftime("%Y%m%d")
    source_meta = []
    for source in SOURCES.values():
        source_meta.append({**{key: value for key, value in source.items() if key != "path"}, "sha256": sha256(source["path"]), "file": source["path"].name})

    write_json("occupations.json", occupations)
    write_json("taxonomies.json", taxonomies)
    write_json("international.json", {
        "module": "international-classifications",
        "isco": {"records": isco, "source": {key: value for key, value in SOURCES["isco"].items() if key != "path"}},
        "esco": {
            "mode": "official-web-service-lookup",
            "selectedVersion": "v1.2.0",
            "language": "en",
            "api": "https://ec.europa.eu/esco/api/search",
            "source": {"id": "esco", "publisher": "European Commission", "vintage": "v1.2.0 web service; v1.2.1 downloadable release available separately", "url": "https://esco.ec.europa.eu/en/use-esco/use-esco-services-api/esco-web-service-api"},
            "integrity": "Live ESCO search results are suggestions from a separate vocabulary; they are not an O*NET-to-ESCO mapping.",
        },
    })
    write_json("manifest.json", {
        "releaseId": release_id,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "mode": "real-source-release",
        "counts": {**counts, **onet_counts},
        "collections": {"occupations": "occupations.json", "taxonomies": "taxonomies.json", "international": "international.json"},
        "sources": source_meta,
        "integrity": {"syntheticRecords": 0, "unresolvedIndustryOccupationCrosswalk": True, "workRhythmData": "not included; source-backed task/activity data only"},
    })
    print(json.dumps({"releaseId": release_id, "counts": {**counts, **onet_counts}}, indent=2))


if __name__ == "__main__":
    main()
